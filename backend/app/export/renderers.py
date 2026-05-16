"""PDF + Excel + JSON + iCal exporters for a Timetable."""
from __future__ import annotations

import io
import json
from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.units import cm

from ..models.domain import Timetable, TimetableRequest


def _columns(req: TimetableRequest) -> list[dict]:
    """Build the interleaved column descriptor list: slots + tea/lunch breaks.

    Each entry is either {"kind":"slot","slot":int} or
    {"kind":"break","label": "TEA BREAK"/"LUNCH BREAK", "duration": int}.
    """
    tc = req.time_config
    cols: list[dict] = []
    for t in range(1, tc.slots_per_day + 1):
        cols.append({"kind": "slot", "slot": t})
        if t == tc.tea_break.after_slot:
            cols.append(
                {"kind": "break", "label": "TEA BREAK", "duration": tc.tea_break.duration_min}
            )
        if t == tc.lunch_break.after_slot:
            cols.append(
                {"kind": "break", "label": "LUNCH BREAK", "duration": tc.lunch_break.duration_min}
            )
    return cols


def _slot_timing(req: TimetableRequest, t: int) -> str:
    timings = req.time_config.slot_timings
    if t - 1 < len(timings):
        return f"{timings[t - 1].start}–{timings[t - 1].end}"
    return ""


# ---------------------------------------------------------------------------
# Grid building
# ---------------------------------------------------------------------------
def _section_grid(tt: Timetable, sec_id: str, days: list[str], slots: list[int]):
    grid: dict[tuple[str, int], list[str]] = defaultdict(list)
    for c in tt.classes:
        if c.section_id != sec_id:
            continue
        cell = grid[(c.day, c.slot)]
        label = c.label or c.course_code
        batch = f" ({c.batch_id})" if c.batch_id else ""
        if label not in cell[0] if cell else True:
            cell.append(f"{label}{batch}")
    return grid


def _faculty_grid(tt: Timetable, faculty_id: str, days: list[str], slots: list[int]):
    grid: dict[tuple[str, int], list[str]] = defaultdict(list)
    for c in tt.classes:
        if c.faculty_id != faculty_id:
            continue
        cell = grid[(c.day, c.slot)]
        text = f"{c.section_id}: {c.label or c.course_code}"
        if text not in cell:
            cell.append(text)
    return grid


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def render_pdf(req: TimetableRequest, tt: Timetable) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        title="Timetable",
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    elements: list = []

    days = req.time_config.days
    slots = list(range(1, req.time_config.slots_per_day + 1))
    cols = _columns(req)
    n_cols = 1 + len(cols)  # +1 for the Day column
    break_col_idxs = [i + 1 for i, c in enumerate(cols) if c["kind"] == "break"]

    # Column widths: Day | slot... | break (wide enough for "BREAK" bold) | ... .
    page_w = landscape(A4)[0] - 2 * (1.2 * cm)
    break_w = 1.5 * cm  # accommodates "BREAK" at 8pt bold with padding
    day_w = 1.7 * cm
    remaining = page_w - day_w - break_w * len(break_col_idxs)
    slot_count = sum(1 for c in cols if c["kind"] == "slot")
    slot_w = max(remaining / slot_count, 1.5 * cm) if slot_count else 2 * cm
    col_widths = [day_w]
    for c in cols:
        col_widths.append(break_w if c["kind"] == "break" else slot_w)

    AMBER = colors.HexColor("#F4C752")
    AMBER_SOFT = colors.HexColor("#FCE9B6")
    NAVY = colors.HexColor("#0E2447")
    INK = colors.HexColor("#161412")

    for sec in req.sections:
        elements.append(
            Paragraph(
                f"<b>{sec.name}</b> &nbsp;·&nbsp; <font color='#6B6356'>{sec.classroom}</font>",
                styles["Title"],
            )
        )
        elements.append(Spacer(1, 4))

        # Header row: Day | (Slot N + timing) ... | BREAK label ...
        header: list = ["Day"]
        for c in cols:
            if c["kind"] == "slot":
                t = c["slot"]
                header.append(f"Slot {t}\n{_slot_timing(req, t)}")
            else:
                header.append(c["label"])

        grid = _section_grid(tt, sec.id, days, slots)
        data = [header]
        for d in days:
            row: list = [d]
            for c in cols:
                if c["kind"] == "break":
                    row.append("")  # body of break column is empty; spanned by header
                    continue
                t = c["slot"]
                cell = grid.get((d, t))
                if cell:
                    row.append("\n".join(cell))
                else:
                    row.append("—" if d == "SAT" else "")
            data.append(row)

        style_cmds: list = [
            # Header band
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            # Body type
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9CCAA")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFFBED"), colors.white]),
            # Day column emphasis
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0E5C9")),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 1), (0, -1), NAVY),
        ]
        # Break columns: amber band, span across all rows, two-line label
        # that fits comfortably inside the column without overflow.
        for bc in break_col_idxs:
            style_cmds.append(("SPAN", (bc, 0), (bc, len(data) - 1)))
            style_cmds.append(("BACKGROUND", (bc, 0), (bc, -1), AMBER))
            style_cmds.append(("TEXTCOLOR", (bc, 0), (bc, -1), INK))
            style_cmds.append(("FONTNAME", (bc, 0), (bc, -1), "Helvetica-Bold"))
            style_cmds.append(("FONTSIZE", (bc, 0), (bc, -1), 8))
            style_cmds.append(("LEADING", (bc, 0), (bc, -1), 10))
            style_cmds.append(("ALIGN", (bc, 0), (bc, -1), "CENTER"))
            style_cmds.append(("VALIGN", (bc, 0), (bc, -1), "MIDDLE"))
            style_cmds.append(("LEFTPADDING", (bc, 0), (bc, -1), 1))
            style_cmds.append(("RIGHTPADDING", (bc, 0), (bc, -1), 1))
            raw = cols[bc - 1]["label"]  # e.g., "TEA BREAK"
            duration = cols[bc - 1]["duration"]
            # Stack as two short lines + duration, all fit in 1.5cm.
            data[0][bc] = raw.replace(" ", "\n") + f"\n{duration} min"

        tbl = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        tbl.setStyle(TableStyle(style_cmds))
        elements.append(tbl)

        # Legend footer per page
        elements.append(Spacer(1, 6))
        elements.append(
            Paragraph(
                "<font size=7 color='#6B6356'>"
                f"Tea break {req.time_config.tea_break.duration_min} min after slot "
                f"{req.time_config.tea_break.after_slot} &nbsp;·&nbsp; "
                f"Lunch break {req.time_config.lunch_break.duration_min} min after slot "
                f"{req.time_config.lunch_break.after_slot} &nbsp;·&nbsp; "
                "Saturday: 1st &amp; 3rd off; locked activities only on others."
                "</font>",
                styles["Normal"],
            )
        )
        elements.append(PageBreak())

    if not elements:
        elements.append(Paragraph("No data", styles["Normal"]))
    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def render_xlsx(req: TimetableRequest, tt: Timetable) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    days = req.time_config.days
    slots = list(range(1, req.time_config.slots_per_day + 1))
    cols = _columns(req)

    header_fill = PatternFill("solid", fgColor="0E2447")
    header_font = Font(bold=True, color="FFFFFF")
    day_fill = PatternFill("solid", fgColor="F0E5C9")
    day_font = Font(bold=True, color="0E2447")
    break_fill = PatternFill("solid", fgColor="F4C752")
    break_font = Font(bold=True, color="161412", size=11)
    body_font = Font(size=9)
    body_alt = PatternFill("solid", fgColor="FFFBED")

    for sec in req.sections:
        ws = wb.create_sheet(title=f"Sec {sec.id}")
        # Column 1 = Day, columns 2.. = slots interleaved with breaks
        ws.cell(row=1, column=1, value="Day")
        for j, c in enumerate(cols, start=2):
            if c["kind"] == "slot":
                t = c["slot"]
                ws.cell(row=1, column=j, value=f"Slot {t}\n{_slot_timing(req, t)}")
            else:
                ws.cell(row=1, column=j, value=c["label"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32

        grid = _section_grid(tt, sec.id, days, slots)
        for i, d in enumerate(days, start=2):
            day_cell = ws.cell(row=i, column=1, value=d)
            day_cell.fill = day_fill
            day_cell.font = day_font
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            for j, c in enumerate(cols, start=2):
                if c["kind"] == "break":
                    ws.cell(row=i, column=j, value="")
                    continue
                t = c["slot"]
                cell_items = grid.get((d, t))
                txt = "\n".join(cell_items) if cell_items else ("—" if d == "SAT" else "")
                cc = ws.cell(row=i, column=j, value=txt)
                cc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                cc.font = body_font
                if i % 2 == 0:
                    cc.fill = body_alt
            ws.row_dimensions[i].height = 46

        # Merge break columns vertically (header row through last data row) so
        # the label sits centred over its column, mirroring the UI grid.
        last_row = 1 + len(days)
        for j, c in enumerate(cols, start=2):
            if c["kind"] != "break":
                continue
            col_letter = ws.cell(row=1, column=j).column_letter
            ws.merge_cells(start_row=1, start_column=j, end_row=last_row, end_column=j)
            merged = ws.cell(row=1, column=j)
            merged.value = c["label"].replace(" ", "\n")
            merged.fill = break_fill
            merged.font = break_font
            merged.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True, text_rotation=90
            )
            ws.column_dimensions[col_letter].width = 5

        # Column widths
        for col in range(1, len(cols) + 2):
            if col == 1:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 8
            else:
                desc = cols[col - 2]
                if desc["kind"] == "slot":
                    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        ws.freeze_panes = "B2"

    # Faculty sheet
    fws = wb.create_sheet(title="Faculty view")
    fws.append(["Faculty", "Day", "Slot", "Section", "Course"])
    for c in tt.classes:
        if c.faculty_id:
            fws.append([c.faculty_id, c.day, c.slot, c.section_id, c.label or c.course_code])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------
def render_json(req: TimetableRequest, tt: Timetable) -> bytes:
    return json.dumps(
        {
            "request": req.model_dump(),
            "timetable": tt.model_dump(),
        },
        default=str,
        indent=2,
    ).encode("utf-8")


# ---------------------------------------------------------------------------
# iCal (per faculty)
# ---------------------------------------------------------------------------
def render_ical(req: TimetableRequest, tt: Timetable, faculty_id: str) -> bytes:
    from icalendar import Calendar, Event

    cal = Calendar()
    cal.add("prodid", "-//Timetable Generator//EN")
    cal.add("version", "2.0")

    # Use next Monday as week start
    today = datetime.now().date()
    monday = today + timedelta(days=(7 - today.weekday()) % 7)
    day_offset = {d: i for i, d in enumerate(["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"])}

    for c in tt.classes:
        if c.faculty_id != faculty_id:
            continue
        timing = (
            req.time_config.slot_timings[c.slot - 1]
            if c.slot - 1 < len(req.time_config.slot_timings)
            else None
        )
        if not timing:
            continue
        d = monday + timedelta(days=day_offset.get(c.day, 0))
        start_h, start_m = [int(x) for x in timing.start.split(":")]
        end_h, end_m = [int(x) for x in timing.end.split(":")]
        ev = Event()
        ev.add("summary", f"{c.label or c.course_code} ({c.section_id})")
        ev.add("dtstart", datetime(d.year, d.month, d.day, start_h, start_m))
        ev.add("dtend", datetime(d.year, d.month, d.day, end_h, end_m))
        ev.add("description", f"Section {c.section_id} | {c.course_code}")
        cal.add_component(ev)

    return bytes(cal.to_ical())
